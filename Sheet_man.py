from openpyxl import load_workbook
from Inventory_obj import Inventory_Object

def check_is_number(string):
    string = string.strip()
    invalid_chars = []
    for char in string:
        if not char.isnumeric():
            invalid_chars.append(char)
    if len(invalid_chars)<1:
        return True, float(string)
    else:
        print(invalid_chars)
        for char in invalid_chars:
            print(string.index(char))
            print(char)
            if char=='-' or char=='.' or char=='$':    
                if char=='-' and string.index(char)!=0:
                    print('neg fail')
                    return False, string
                elif char=='.' and string.index(char)!=len(string)-3:
                    print(char)
                    print('period fail')
                    return False, string
                elif char=='$':
                    if string.index(char)!=0 or string.index(char)!=len(string)-1:
                        print(('$ fail'))
                        return False, string
            else:
                print('generic fail')
                return False, string
        return True, float(string)
            
            


class Sheet_Man:
    def __init__(self, path, man, block_ids):
        self.book = None
        self.sheet = None
        self.path = None
        self.block_ids = block_ids
        self.manager = man
        self.open_sheet(path)
        self.inventory_length = self.get_inv_length()

    def open_sheet(self, path=None):
        if path==None: path=self.path
        try:
            self.book = load_workbook(path)
            sheet_name = self.book.sheetnames[0]
            self.sheet = self.book[sheet_name]
            self.path = path
        except:
            print('failed to load excel sheet at path')
    def get_inv_length(self):
        bounce=False
        count = int(self.block_ids['start_row'])
        while bounce==False:
            count+=1
            position = self.block_ids['start_column']+str(count)
            if self.sheet[position].value==None:
                bounce=True
        return count

    def get_list(self, config):
        update_list = []
        for x in range(int(self.block_ids['start_row']), self.inventory_length):
            href = self.sheet[self.block_ids['href']+str(x)].value
            #checks if it needs updates
            if self.sheet[self.block_ids['last_update']+str(x)].value==None:
                update_list.append(self.manager.build_inv_obj(href, str(x)))
        return update_list

    def update_block(self, item, key, value):
        validation, rvalue = check_is_number(value)
        print('val: {}\nrval: {}'.format(validation,rvalue))
        if validation:
            self.sheet[str(self.block_ids[key])+str(item.row)] = float(rvalue)
        else:
            self.sheet[str(self.block_ids[key])+str(item.row)] = rvalue
        self.book.save(self.path)
        print('updated book and saved')
